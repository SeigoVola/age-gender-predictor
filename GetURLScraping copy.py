from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
import random
from openpyxl import Workbook, load_workbook
import os
from urllib.parse import urljoin
from selenium.common.exceptions import TimeoutException

# ランダム待機（ボット検出回避・負荷軽減）
def random_sleep(min_seconds: float = 2.0, max_seconds: float = 6.0) -> None:
    sleep(random.uniform(min_seconds, max_seconds))

# Chrome設定
options = webdriver.ChromeOptions()
# options.add_argument('--headless')
driver = webdriver.Chrome(options=options)

# セッション情報をリセット
# try:
#     driver.delete_all_cookies()
#     driver.get("about:blank")
#     driver.execute_script("window.sessionStorage.clear(); window.localStorage.clear();")
#     print("セッション情報をリセットしました")
# except Exception as e:
#     print(f"セッションリセット時にエラー: {e}")





# driver.get(target_url)

wait = WebDriverWait(driver, 10)

url_data = []
comment_data = []

# ＝＝＝商品URLの取得＝＝＝

# # # ＝＝＝サイドバーのカテゴリーボタンを取得＝＝＝
# link_list = wait.until(EC.presence_of_element_located((
#     By.XPATH, "//*[@id='__next']/div[2]/div/main/div[3]/div/div[2]/nav/section[1]/section[1]/ul"
# )))
# link_elems = link_list.find_elements(By.XPATH, ".//a[@href]")
# # print(link_list)
# raw_urls = [a.get_attribute("href") for a in link_elems]
# urls = [urljoin(driver.current_url, u) for u in raw_urls if u]              # 絶対URL化
# unique_urls = list(dict.fromkeys(urls))      
# # print(unique_urls)


# # # ＝＝＝商品ページのスクレイピング＝＝＝
# for url in unique_urls:
#     driver.get(url)
#     WebDriverWait(driver, 30).until(lambda d: d.execute_script("return document.readyState") == "complete")
#     random_sleep(3, 7)


# Excelファイルにコメント・性別・年代を書き込む関数
def write_to_excel(comment, gender, age):
    file_name = "DataSet.xlsx"
    sheet_name = "Sheet1"

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
    else:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = sheet_name
        ws.append(["Comment", "Gender", "Age"])  # ヘッダー追加

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=comment)
    ws.cell(row=next_row, column=2, value=gender)
    ws.cell(row=next_row, column=3, value=age)

    wb.save(file_name)

# =====　次のページへ移行 =====

# Pnoでページ番号を指定してアクセス（存在しなくなるまで繰り返す）
Pno = 1  # 開始ページ
while True:
    first_url = "https://www.croxyproxy.com/"
    driver.get(first_url)

    # target_url = "https://zozo.jp/"
    # 練習用
    target_url = "https://zozo.jp/category/tops/"
    page_url = f"{target_url}?pno={Pno}"
    print(f"ページ遷移: {page_url}")
    # 指定のXPathの入力欄にpage_urlを入力し、ボタンをクリック
    try:
        input_elem = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((
            By.XPATH, "/html/body/div[2]/div/div/div/div[2]/div[2]/form/div/div/div/div/input"
        )))
        input_elem.clear()
        input_elem.send_keys(page_url)

        button_elem = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((
            By.XPATH, "/html/body/div[2]/div/div/div/div[2]/div[2]/form/div/span/button"
        )))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button_elem)
        sleep(1)
        button_elem.click()
        WebDriverWait(driver, 15).until(lambda d: d.execute_script("return document.readyState") == "complete")
        random_sleep(2, 5)
    except Exception as e:
        print(f"入力/クリック処理でエラー: {e}")

    # 対象コンテナの存在確認（なければ終了）
    try:
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.XPATH, "/html/body/div[1]/div[2]/div/main/div/div[2]/div[1]/div[2]"
        )))
    except TimeoutException:
        print(f"ページが存在しませんでした。終了します。（pno={Pno}）")
        
        break

    

    # ここでページ内のスクレイピング処理を行う
    # （既存の商品取得ロジックをこのループ内に移動すると、各ページで実行されます）
    # ===== 商品ページの取得 =====
    # 方法1: 親要素を取得してから子要素を検索
    merchandise_list = wait.until(EC.presence_of_element_located((
        By.XPATH, "/html/body/div[1]/div[2]/div/main/div/div[2]/div[1]/div[2]"
    )))
    # print("親要素が見つかりました:", merchandise_list.tag_name)
    merchandise_elems = merchandise_list.find_elements(By.XPATH, "./div/div[1]/a")
    # print(f"2つ下の階層のdiv要素数: {len(merchandise_elems)}")
    raw_merchandise_urls = [a.get_attribute("href") for a in merchandise_elems]
    # print(f"URL数: {len(raw_merchandise_urls)}")
    merchandise_urls = [urljoin(driver.current_url, u) for u in raw_merchandise_urls if u]              # 絶対URL化
    if len(merchandise_urls) >= 10:
        merchandise_urls = merchandise_urls[5:-5]
    else:
        merchandise_urls = []
    unique_merchandise_urls = list(dict.fromkeys(merchandise_urls))      
    print(f"URL数: {len(unique_merchandise_urls)}")
    # print(unique_merchandise_urls)

    # 収集したURLをurl_dataに追加
    url_data.extend(unique_merchandise_urls)
    print(url_data)


    random_sleep(3, 6)
    driver.quit()
        
   # 次のページへ
    Pno += 1
# ＝＝＝ スクレイピングを終了 ＝＝＝




# # zozotownのスクレイピング




