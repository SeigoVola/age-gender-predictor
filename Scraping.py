from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
import random
from openpyxl import Workbook, load_workbook
import os
from urllib.parse import urljoin
from selenium.common.exceptions import TimeoutException

# ランダム待機（ボット検出回避・負荷軽減）
def random_sleep(min_seconds: float = 2.0, max_seconds: float = 6.0) -> None:
    sleep(random.uniform(min_seconds, max_seconds))

# Chrome設定
options = webdriver.ChromeOptions()
# options.add_argument('--headless')
driver = webdriver.Chrome(options=options)

# セッション情報をリセット
try:
    driver.delete_all_cookies()
    driver.get("about:blank")
    driver.execute_script("window.sessionStorage.clear(); window.localStorage.clear();")
    print("セッション情報をリセットしました")
except Exception as e:
    print(f"セッションリセット時にエラー: {e}")
# first_url = "https://www.croxyproxy.com/"
# driver.get(first_url)

# target_url = "https://zozo.jp/"
# 練習用
target_url = "https://zozo.jp/category/tops/"




# driver.get(target_url)

wait = WebDriverWait(driver, 10)

url_data = []
comment_data = []

# ＝＝＝商品URLの取得＝＝＝

# # # ＝＝＝サイドバーのカテゴリーボタンを取得＝＝＝
# link_list = wait.until(EC.presence_of_element_located((
#     By.XPATH, "//*[@id='__next']/div[2]/div/main/div[3]/div/div[2]/nav/section[1]/section[1]/ul"
# )))
# link_elems = link_list.find_elements(By.XPATH, ".//a[@href]")
# # print(link_list)
# raw_urls = [a.get_attribute("href") for a in link_elems]
# urls = [urljoin(driver.current_url, u) for u in raw_urls if u]              # 絶対URL化
# unique_urls = list(dict.fromkeys(urls))      
# # print(unique_urls)


# # # ＝＝＝商品ページのスクレイピング＝＝＝
# for url in unique_urls:
#     driver.get(url)
#     WebDriverWait(driver, 30).until(lambda d: d.execute_script("return document.readyState") == "complete")
#     random_sleep(3, 7)


# Excelファイルにコメント・性別・年代を書き込む関数
def write_to_excel(comment, gender, age):
    file_name = "DataSet.xlsx"
    sheet_name = "Sheet1"

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
    else:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = sheet_name
        ws.append(["Comment", "Gender", "Age"])  # ヘッダー追加

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=comment)
    ws.cell(row=next_row, column=2, value=gender)
    ws.cell(row=next_row, column=3, value=age)

    wb.save(file_name)

# =====　次のページへ移行 =====

# Pnoでページ番号を指定してアクセス（存在しなくなるまで繰り返す）
Pno = 1  # 開始ページ

# =====　商品ページへのアクセス =====

# ExcelからURLリストを読み込み（同フォルダの「データ.xlsx」、A列）
excel_path = os.path.join(os.path.dirname(__file__), "データ.xlsx")
wb_urls = load_workbook(excel_path)
ws_urls = wb_urls.active
unique_merchandise_urls = [str(c.value).strip() for c in ws_urls['A'] if c.value]
# 重複を順序維持で除去
unique_merchandise_urls = list(dict.fromkeys(unique_merchandise_urls))
for url in unique_merchandise_urls:
    # print("動いている")
    driver.get(url)
    scroll_amount = 1000 * 1.8
    WebDriverWait(driver, 30).until(lambda d: d.execute_script("return document.readyState") == "complete")
    print("動いている")
    random_sleep(4, 8)

    # ===== レビュータブをクリック =====
    search_button = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "a.c-tab-item.reviewTab"))
    )
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", search_button)
    sleep(1)
    ActionChains(driver).move_to_element(search_button).click().perform()

    sleep(5)

    # 投稿件数を取得
    post_count = wait.until(EC.presence_of_element_located(
        (By.XPATH, "/html/body/div[1]/div[3]/div[2]/article/div/div/div/div/div[1]/div[2]/div[3]/div[1]/div/a[3]/div/div/p")
    ))
    # 全角括弧を除去してから整数変換
    text = post_count.text.split('件')[0].replace('（', '').replace('）', '')
    post_number = int(text) if text.isdigit() else 0
    print(f"レビュー件数: {post_number}")

    # 件数に応じた分岐
    use_modal = False
    if post_number == 0:
        print("レビュー0件のためスキップします")
        continue
    elif 1 <= post_number <= 3:
        print("レビュー1-3件のため簡易処理（モーダル未使用）")
        use_modal = False
    else:
        print("レビュー4件以上のためモーダルを開いて処理します")
        # 「すべてのレビューを見る」ボタンをクリック
        review_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(@class, 'button--kOWzLwXw isPC--1Wt6kFir')]")
        ))
        review_button.click()
        sleep(5)

        # スクロール対象のモーダル要素取得
        scrollable_modal = wait.until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="zozotown"]/div[13]/div/div/div/div[1]/div[2]'))
        )

        # コメント1つ分の高さ取得（概算スクロール用）
        comment_element = driver.find_element(By.XPATH, '//*[@id="zozotown"]/div[13]/div/div/div/div[1]/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/div/div/div[1]/div')
        comment_height = comment_element.size['height']
        scroll_amount = comment_height * 1.8
        use_modal = True

    # ===== レビュー取得とスクロールループ =====
    try:
        while len(comment_data) < post_number:
            try:
                # コメント・性別・年代の要素を取得
                comment_elems = driver.find_elements(By.XPATH, "//*[@id='zozotown']/div[13]/div/div/div/div[1]/div[2]/div[2]/div[2]/div/div/div[1]/div//div[contains(@class,'content--2bDyVHrb')]")
                gender_elems = driver.find_elements(By.XPATH, "//*[@id='zozotown']/div[13]/div/div/div/div[1]/div[2]/div[2]/div[2]/div/div/div[1]/div//div[contains(@class,'questions--2g62oqc-')]")
                sleep(5)

                for i in range(min(len(comment_elems), len(gender_elems))):
                    comment_text = comment_elems[i].text
                    gender_text = gender_elems[i].text
                    sleep(5)

                    # 重複チェック
                    if comment_text in comment_data:
                        continue
                    comment_data.append(comment_text)

                    # 性別の判定
                    if "男性" in gender_text:
                        gender = "male"
                    elif "女性" in gender_text:
                        gender = "female"
                    else:
                        gender = "null"

                    # 年代の判定
                    if "〜18歳" in gender_text:
                        age = "10代"
                    elif "19〜22歳" in gender_text or "23〜29歳" in gender_text:
                        age = "20代"
                    elif "30〜34歳" in gender_text or "35〜39歳" in gender_text:
                        age = "30代"
                    elif "40〜44歳" in gender_text or "45〜49歳" in gender_text:
                        age = "40代"
                    elif "50歳〜" in gender_text:
                        age = "50代"
                    else:
                        age = "null"

                    # Excel に保存
                    write_to_excel(comment_text, gender, age)

                    sleep(1)

                # モーダル内スクロール
                driver.execute_script("arguments[0].scrollTop += arguments[1]", scrollable_modal, scroll_amount)
                print("スクロール完了")
                sleep(5)

            except Exception as e:
                print(f"スクロール・取得中にエラー: {e}")
                break

    except Exception as e:
        print(f"全体処理エラー: {e}")
# ＝＝＝ スクレイピングを終了 ＝＝＝
random_sleep(3, 6)
driver.quit()



# # zozotownのスクレイピング




